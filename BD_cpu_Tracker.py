''' This program is used to track daily AMD cpu prices from startech bd and Ryans Computer and write it to an existing excel spreadsheet '''
# python 3.8.2
import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

url = ['https://www.startech.com.bd/component/processor/amd-processor?limit=90',
       'https://www.ryanscomputers.com/grid/desktop-component-processor?query=10-%40%23%40%7C']
agent = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.61 Safari/537.36'}

content = requests.get(url[0], headers=agent).text
soup = BeautifulSoup(content, 'lxml')

# empty list is created to loop through later
star_price = []
star_product = []

for product_class in soup.find_all(
        'div', class_='col-xs-12 col-md-4 product-layout grid'):
    product_name = product_class.find(
        'div', class_='product-content-blcok').h4.a.text
    # list is populated through append method
    star_product.append(product_name)
    price1 = product_class.find('div', class_='price space-between')
    price = price1.find('span').text[:-1]
    price = price.replace(',', '')  # for number formatting comma is removed
    star_price.append(price)

ryans_product = []
ryans_price = []

content = requests.get(url[1], headers=agent).text
soup = BeautifulSoup(content, 'lxml')

for li_class in soup.find_all('li', class_='col-sm-4 col-md-4 col-lg-4'):
    title2 = li_class.find('div', class_='product-content-info').a.text
    title2 = title2.split('.')[0]
    title = title2[:-2]
    ryans_product.append(title)
    price = li_class.find('span', class_='price').text
    price = price.strip().replace(',', '')
    ryans_price.append(price)

# this file need to be in the directory
wb = load_workbook(filename='sunday_update.xlsx')
sheet = wb.active
max_c = sheet.max_column
max_r = sheet.max_row

# complete new file can be added via wb = Workbook(filename='sunday_update.xlsx')

time = datetime.datetime.now()
date = time.strftime('%d' + '-' + '%b')  # fromat date as date-month

# the date this tracker records the prices
sheet.cell(column=max_c+1, row=2, value=date)

for x in range(0, len(star_price)):
    # earlier comma is removed to make it integer here
    sheet.cell(column=max_c+1, row=x+3, value=int(star_price[x]))
    x = +1

for x in range(0, len(ryans_price)):
    sheet.cell(column=max_c+1, row=x+28, value=int(ryans_price[x]))
    x = +1
wb.save(filename='easy2.xlsx')  # file will be saved as a new file

print('A-OK')

# to check if any new cpu has been added
# 12 and 25 is manually counted and added as a constant
if len(ryans_product) == 12:
    print('No new products in Ryans')
if len(star_product) == 25:
    print('All good in Startech too :D')

# if new products are added then
# sheet.cell(column=3, row=x+3, value=star_product[x]) this code need to be added

# applied knowledge from stack exchange, youtube: corey schaffer, openpyxl documentation
# VSCode is awesome
